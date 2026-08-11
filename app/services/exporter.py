import io
import json
import os
import re
import ast
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

TRANSCRIPTS_DIR = "./exported_transcripts"
os.makedirs(TRANSCRIPTS_DIR, exist_ok=True)


def remove_in_text_phrase_loops(text: str) -> str:
    """
    Cleans internal string phrase repetition loops (e.g. 'I don't know how to do it' repeated 30 times).
    """
    if not text or len(text) < 8:
        return text

    parts = re.split(r'(?<=[.!?,\n])\s+', text)
    if len(parts) <= 1:
        return text

    clean_parts = []
    seen = {}
    prev_norm = ""

    for part in parts:
        p_strip = part.strip()
        if not p_strip:
            continue
        
        norm = re.sub(r'[^\w\s]', '', p_strip.lower()).strip()
        if not norm:
            continue

        cnt = seen.get(norm, 0) + 1
        seen[norm] = cnt

        if cnt > 2 or norm == prev_norm:
            continue

        prev_norm = norm
        clean_parts.append(p_strip)

    result = " ".join(clean_parts).strip()
    return result if result else text


def clean_text_segment(text: str) -> str:
    """Timestamp prefixes jaise [0.3s - 31.0s] aur internal repeated phrase loops ko clean karta hai."""
    txt = re.sub(r'\[\s*\d+(\.\d+)?s\s*-\s*\d+(\.\d+)?s\s*\]', '', text).strip()
    return remove_in_text_phrase_loops(txt)


def parse_dialogue_text_into_turns(text: str) -> list:
    """
    Sirf tab chalega jab raw text string di gayi ho. 
    Yeh sentences ko tod kar turns banata hai.
    """
    clean_txt = clean_text_segment(text)
    raw_sentences = re.split(r'(?<=[.!?])\s+', clean_txt)
    raw_sentences = [s.strip() for s in raw_sentences if s.strip()]

    turns = []
    for idx, sentence in enumerate(raw_sentences):
        # No structured speaker/diarization data is available at all in this
        # path (it only runs on a bare merged text string) - use content
        # keywords to guess, and mark genuinely ambiguous lines as Unknown
        # rather than assuming a fixed speaking order.
        lower_s = sentence.lower()
        if "thank you for calling" in lower_s or "how can i assist" in lower_s or "let me check" in lower_s or "i apologize" in lower_s or "anything else" in lower_s:
            role = "Receptionist"
        elif "took you long enough" in lower_s or "unacceptable" in lower_s or "my order" in lower_s or "i need help" in lower_s:
            role = "Customer"
        else:
            role = "Unknown"

        turns.append({
            "speaker": "",
            "role": role,
            "message": sentence
        })
    return turns


def parse_and_format_transcript(transcript_data) -> list:
    """
    Data ko baghair kisi ghalat assumption ke safely list of dicts mein convert karta hai.
    """
    if not transcript_data or transcript_data in ["null", "N/A", {}]:
        return []

    if isinstance(transcript_data, str):
        try:
            transcript_data = json.loads(transcript_data)
        except Exception:
            try:
                transcript_data = ast.literal_eval(transcript_data)
            except Exception:
                pass

    if isinstance(transcript_data, list):
        formatted = []
        for turn in transcript_data:
            if isinstance(turn, dict):
                msg = turn.get("message") or turn.get("text") or turn.get("content") or ""
                cleaned_msg = clean_text_segment(str(msg))
                if cleaned_msg:
                    entry = {"message": cleaned_msg}
                    # IMPORTANT: preserve the speaker/role the AI engine already
                    # resolved via diarization + LLM. Previously this field was
                    # dropped here, which is exactly what caused every line to
                    # fall back to a hardcoded default later in the pipeline.
                    if turn.get("speaker"):
                        entry["speaker"] = turn.get("speaker")
                    if turn.get("role"):
                        entry["role"] = turn.get("role")
                    formatted.append(entry)
            else:
                formatted.extend(parse_dialogue_text_into_turns(str(turn)))
        return formatted

    if isinstance(transcript_data, dict):
        if "conversation" in transcript_data and isinstance(transcript_data["conversation"], list):
            return parse_and_format_transcript(transcript_data["conversation"])
        if "dialogues" in transcript_data and isinstance(transcript_data["dialogues"], list):
            return parse_and_format_transcript(transcript_data["dialogues"])
        if "transcript" in transcript_data:
            return parse_and_format_transcript(transcript_data["transcript"])
        if "full_text" in transcript_data and isinstance(transcript_data["full_text"], str):
            return parse_dialogue_text_into_turns(transcript_data["full_text"])

    return parse_dialogue_text_into_turns(str(transcript_data))


def fix_transcript_roles_with_ollama(raw_conversation_list):
    """
    Merged strings ya raw list ko sentences mein split kar ke 
    smart rule-based tareeqay se Speaker 1 (Receptionist) aur Speaker 2 (Customer) assign karta hai,
    aur repetitive greetings (e.g. 70 repeated 'Hello's) ko deduplicate karta hai.
    """
    if not raw_conversation_list:
        return []

    # Check if raw_conversation_list already contains pre-formatted turn dictionaries
    if isinstance(raw_conversation_list, list) and all(isinstance(x, dict) for x in raw_conversation_list):
        deduped = []
        greeting_cnt = 0
        prev_msg = ""

        def _normalize_role(value):
            """Maps whatever role/speaker terminology upstream used (Agent,
            Support Agent, SPEAKER_00, Caller, etc.) into this report's
            Receptionist/Customer labels. Returns None if it can't tell."""
            v = str(value).lower()
            if "agent" in v or "receptionist" in v or "support" in v:
                return "Receptionist"
            if "customer" in v or "caller" in v or "client" in v:
                return "Customer"
            return None

        for item in raw_conversation_list:
            msg = str(item.get("message") or item.get("text") or "").strip()
            if not msg:
                continue
            clean_msg = clean_text_segment(msg)
            lower_msg = clean_msg.lower().strip()
            if not lower_msg:
                continue
            
            if lower_msg in ["hello", "hello.", "hi", "hi."]:
                greeting_cnt += 1
                if greeting_cnt > 2:
                    continue
            if lower_msg == prev_msg:
                continue
            prev_msg = lower_msg
            
            spk = item.get("speaker") or item.get("speaker_label", "")
            # Only assign role based on speaker identity, not fixed pattern
            role = item.get("role")
            # If no explicit role, try to normalize whatever the speaker field says
            if not role:
                role = _normalize_role(spk)
            else:
                role = _normalize_role(role) or role

            if not role:
                # We genuinely have no speaker info for this line - mark it
                # honestly as Unknown instead of defaulting every unresolved
                # line to "Receptionist", which was silently mislabeling
                # entire calls as 100% one speaker.
                role = "Unknown"
            
            deduped.append({"speaker": spk, "role": role, "message": clean_msg})
            
        if deduped:
            return deduped

    # Updated deduplication: only remove consecutive duplicate messages from the same speaker
    deduped = []
    prev_msg = ""
    prev_speaker = None
    greeting_cnt = 0
    
    for item in raw_conversation_list:
        msg = str(item.get("message") or item.get("text") or "").strip()
        if not msg:
            continue
        clean_msg = clean_text_segment(msg)
        lower_msg = clean_msg.lower().strip()
        if not lower_msg:
            continue
        
        # greeting deduplication (allow up to 2 greetings overall)
        if lower_msg in ["hello", "hello.", "hi", "hi."]:
            greeting_cnt += 1
            if greeting_cnt > 2:
                continue
        
        # avoid dropping the same utterance if spoken by a different speaker
        spk = item.get("speaker") or item.get("speaker_label", "")
        if lower_msg == prev_msg and spk == prev_speaker:
            continue
        
        prev_msg = lower_msg
        prev_speaker = spk
        
        # role normalization
        role = item.get("role")
        if not role:
            role = _normalize_role(spk)
        else:
            role = _normalize_role(role) or role
        
        if not role:
            role = "Unknown"
        
        deduped.append({"speaker": spk, "role": role, "message": clean_msg})
    
    if deduped:
        return deduped
    full_text = " ".join([item.get("message", "") if isinstance(item, dict) else str(item) for item in raw_conversation_list])
    
    if not full_text.strip():
        return []

    # 2. Sentences mein todne ke liye regex use karein (. ! ? ke baad space)
    sentences = re.split(r'(?<=[.!?])\s+', full_text)
    sentences = [s.strip() for s in sentences if s.strip()]

    # Deduplicate repetitive greetings & consecutive duplicates
    deduped_sentences = []
    prev_s = ""
    greeting_count = 0
    for s in sentences:
        clean_s = clean_text_segment(s)
        lower_s = clean_s.lower().strip()
        if not lower_s:
            continue
            
        if lower_s in ["hello", "hello.", "hi", "hi."]:
            greeting_count += 1
            if greeting_count > 2:
                continue
        if lower_s == prev_s:
            continue
        prev_s = lower_s
        deduped_sentences.append(clean_s)

    sentences = deduped_sentences

    # 3. Rules ke mutabiq roles assign karein
    corrected_turns = []
    
    for idx, sentence in enumerate(sentences):
        is_receptionist = False
        lower_s = sentence.lower()
        
    # Updated role assignment heuristics with expanded keyword lists
    for idx, sentence in enumerate(sentences):
        lower_s = sentence.lower()
        is_receptionist = False
        
        # Receptionist cues
        receptionist_keywords = [
            "thank you for calling", "how can i assist", "let me check",
            "i apologize", "anything else", "good morning", "good afternoon",
            "please hold", "please wait", "may i have", "may i know"
        ]
        # Customer cues
        customer_keywords = [
            "i need help", "i have an issue", "i want", "i'm calling", "my order", "my loan",
            "i don't understand", "i don't know", "i'm not happy", "unacceptable",
            "took you long enough", "worst", "can you", "please"
        ]
        
        if any(kw in lower_s for kw in receptionist_keywords):
            is_receptionist = True
        elif any(kw in lower_s for kw in customer_keywords):
            is_receptionist = False
        else:
            # No clear cue – mark as Unknown
            corrected_turns.append({"speaker": "", "role": "Unknown", "message": sentence})
            continue
        
        speaker = "Speaker 1" if is_receptionist else "Speaker 2"
        role = "Receptionist" if is_receptionist else "Customer"
        
        corrected_turns.append({"speaker": speaker, "role": role, "message": sentence})

        if "thank you for calling" in lower_s or "how can i assist" in lower_s or "let me check" in lower_s or "i apologize" in lower_s or "anything else" in lower_s:
            is_receptionist = True
        elif "took you long enough" in lower_s or "unacceptable" in lower_s or "worst" in lower_s:
            is_receptionist = False
        else:
            # No content signal either way - mark honestly as Unknown instead
            # of guessing, since this raw-text path has no diarization data.
            corrected_turns.append({
                "speaker": "",
                "role": "Unknown",
                "message": sentence
            })
            continue

        speaker = "Speaker 1" if is_receptionist else "Speaker 2"
        role = "Receptionist" if is_receptionist else "Customer"

        corrected_turns.append({
            "speaker": speaker,
            "role": role,
            "message": sentence
        })

    return corrected_turns



def save_transcript_to_json_file(call_id, raw_transcript) -> tuple[str, str]:
    """
    Structured transcript ko JSON file mein save karta hai.
    """
    formatted_dialogues = parse_and_format_transcript(raw_transcript)

    if not formatted_dialogues:
        return "N/A", ""

    formatted_dialogues = fix_transcript_roles_with_ollama(formatted_dialogues)

    file_name = f"transcript_{call_id}.json"
    abs_path = os.path.abspath(os.path.join(TRANSCRIPTS_DIR, file_name))

    output_data = {
        "call_id": str(call_id),
        "total_turns": len(formatted_dialogues),
        "conversation": formatted_dialogues
    }

    try:
        with open(abs_path, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=4, ensure_ascii=False)

        file_url = abs_path.replace("\\", "/")
        if not file_url.startswith("/"):
            file_url = "/" + file_url
            
        return abs_path, f"file://{file_url}"
    except Exception as e:
        print(f"Error saving transcript JSON file: {e}")
        return "Error Saving File", ""


def extract_val(item, key, default="N/A"):
    if key == "summary":
        val = getattr(item, "summary", None)
        if val not in [None, "", "null"]:
            return val
    if key == "transcript":
        val = getattr(item, "transcript_json", None)
        if val not in [None, "", "null", {}]:
            return val

    val = getattr(item, key, None)
    if val not in [None, "", "null"]:
        return val

    if isinstance(item, dict):
        if key in item and item[key] not in [None, "", "null"]:
            return item[key]
        if key == "transcript" and "transcript_json" in item:
            return item["transcript_json"]

    return default


def generate_excel_report(calls):
    """Excel generator with HYPERLINK."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Analytics"
    ws.views.sheetView[0].showGridLines = True

    HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="34D399")
    DATA_FONT = Font(name="Segoe UI", size=10, color="1E293B")
    LINK_FONT = Font(name="Segoe UI", size=10, color="2563EB", underline="single")
    
    BORDER_THIN = Side(border_style="thin", color="CBD5E1")
    CELL_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)

    headers = [
        "Customer Name", "Agent / Dept", "Date & Time", "Duration", 
        "Category", "Problem Statement", "Solution", "Summary", "Resolved", 
        "Sentiment", "Status", "Transcript File"
    ]
    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, call in enumerate(calls, start=2):
        created_at = extract_val(call, "created_at", None)
        date_str = created_at.strftime("%Y-%m-%d %H:%M") if hasattr(created_at, 'strftime') else "N/A"
        
        call_id = extract_val(call, "id", f"row_{row_idx}")
        raw_transcript = extract_val(call, "transcript", None)
        
        abs_path, file_url = save_transcript_to_json_file(call_id, raw_transcript)

        row_values = [
            extract_val(call, 'customer_name', "Unknown"),
            extract_val(call, 'agent_name', "Support AI"),
            date_str,
            extract_val(call, 'duration', "02:45"),
            extract_val(call, 'issue_category', "General"),
            extract_val(call, 'problem_statement', "N/A"),
            extract_val(call, 'solution', "N/A"),
            extract_val(call, 'summary', "N/A"),
            extract_val(call, 'resolved', "NO"),
            extract_val(call, 'sentiment', "Neutral"),
            extract_val(call, 'status', "PROCESSING"),
            f'={file_url}' if file_url else "N/A"
        ]
        
        for col_idx, val in enumerate(row_values, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = CELL_BORDER
            c.alignment = Alignment(
                horizontal="center" if col_idx in [3, 4, 5, 9, 10, 11] else "left", 
                vertical="top",
                wrap_text=True
            )

            if col_idx == 12 and file_url:
                c.value = f'=HYPERLINK("{file_url}", "Open transcript_{call_id}.json")'
                c.font = LINK_FONT
            else:
                c.value = str(val) if val is not None else "N/A"
                c.font = DATA_FONT

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        if col_letter in ["F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 40
        elif col_letter == "L":
            ws.column_dimensions[col_letter].width = 32
        else:
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 30)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream


def generate_pdf_report(calls):
    """PDF Report Generator."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=16, leading=20, textColor=colors.HexColor('#0f172a'), spaceAfter=4)
    meta_style = ParagraphStyle('DocMeta', parent=styles['Normal'], fontSize=10, leading=14, textColor=colors.HexColor('#475569'), spaceAfter=12)
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=9, leading=13, textColor=colors.HexColor('#1e293b'))

    story.append(Paragraph("<b>AI Call Intelligence Platform - Summary Report</b>", title_style))
    story.append(Paragraph(f"Total Calls Exported: {len(calls)}", meta_style))

    for idx, call in enumerate(calls, 1):
        customer = extract_val(call, 'customer_name', 'Unknown')
        agent = extract_val(call, 'agent_name', 'Support AI')
        duration = extract_val(call, 'duration', '02:45')
        category = extract_val(call, 'issue_category', 'General')
        problem = extract_val(call, 'problem_statement', 'N/A')
        solution = extract_val(call, 'solution', 'N/A')
        summary = extract_val(call, 'summary', 'N/A')
        resolved = extract_val(call, 'resolved', 'NO')
        sentiment = extract_val(call, 'sentiment', 'Neutral')
        status = extract_val(call, 'status', 'COMPLETED')
        
        call_id = extract_val(call, "id", f"row_{idx}")
        raw_transcript = extract_val(call, "transcript", None)
        abs_path, file_url = save_transcript_to_json_file(call_id, raw_transcript)

        link_html = f'<a href="{file_url}" color="blue"><u>Open {os.path.basename(abs_path)}</u></a>' if file_url else "N/A"

        card_data = [
            [Paragraph(f"<b>#{idx} Customer:</b> {customer}", cell_style), Paragraph(f"<b>Agent/Dept:</b> {agent}", cell_style)],
            [Paragraph(f"<b>Category:</b> {category}", cell_style), Paragraph(f"<b>Duration:</b> {duration}", cell_style)],
            [Paragraph(f"<b>Resolved:</b> {resolved}", cell_style), Paragraph(f"<b>Sentiment:</b> {sentiment} | <b>Status:</b> {status}", cell_style)],
            [Paragraph(f"<b>Problem Statement:</b> {problem}", cell_style), ""],
            [Paragraph(f"<b>Solution:</b> {solution}", cell_style), ""],
            [Paragraph(f"<b>Executive Summary:</b> {summary}", cell_style), ""],
            [Paragraph(f"<b>Transcript JSON Path:</b> {link_html}", cell_style), ""]
        ]

        table = Table(card_data, colWidths=[270, 270])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('SPAN', (0, 3), (1, 3)),
            ('SPAN', (0, 4), (1, 4)),
            ('SPAN', (0, 5), (1, 5)),
            ('SPAN', (0, 6), (1, 6)),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 10))

    doc.build(story)
    buffer.seek(0)
    return buffer